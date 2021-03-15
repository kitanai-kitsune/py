import openpyxl
from easygui import *

filename = fileopenbox("ファイルを開く")

book = openpyxl.load_workbook(filename)

sht = book.get_sheet_by_name("Sheet1")
sht2 = book.get_sheet_by_name("Sheet2")

colV = sht["V"]

targets = [] #需要进行排序的所有数据

for row in sht.iter_rows(min_row=2):
    if row[21].value is None:
        continue
    values = []
    for col in row:
        values.append(col.value)

    print(values)
    targets.append(values)

print(f"共有{len(targets)}条数据需要排列")

#把需要排序的数据复制进第二张sheet
for data in targets:
    sht2.append(data)

#把需要排序的数据从第一张sheet删除
for col in colV:
    if col.value is not None:
        RowNumber = col.row
        sht.delete_rows(RowNumber)
# 以上 把第一页中需要排列的数据剪切到第二页

print(sht2["H1"].value)

for i in range(sht2.max_row):

    colH = sht["H"]
    same_busho = 0
    same_busho_kojin = 0

    compare_list = []
    old_list = []

    k = 0

    for col in colH:
        if col.value == sht2["H1"].value:
            same_busho += 1 #同部署的邮箱数量（包括共用等）

        if col.value == sht2["H1"].value and "mbox" not in sht["A" + str(col.row)].value and "atenda" not in sht["A" + str(col.row)].value\
                and sht["A" + str(col.row)].value.split("-")[0][-1].isdigit() == True: #同部署的个人邮箱
            same_busho_kojin += 1 #同部署的个人邮箱数量

    print(f"同部署的共有{same_busho}条,不同部署的共有{sht.max_row - same_busho}条")
    print(f"同部署的个人共有{same_busho_kojin}条")

    if same_busho != 0:  #如果同部署的邮箱数量不为0
        print("进入1号分支")

        for col in colH:
            if col.value == sht2["H1"].value and "mbox" not in sht["A" + str(col.row)].value and "atenda" not in sht["A" + str(col.row)].value\
                    and sht["A" + str(col.row)].value.split("-")[0][-1].isdigit() == True: #同部署的个人邮箱
                lastone = col.row

                if sht2["J1"].value is not None:#（并且不是空值的情况）

                    #センター的五个职位的判断
                    if "支店長" in sht2["J1"].value: #（支店長的情况）
                        if sht["J" + str(col.row)].value is not None and "支店長" in sht["J" + str(col.row)].value: #（如果这边也有支店長）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "支店長" not in sht["J" + str(col.row)].value: #（如果这边有职位 但不是支店长）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    elif "所長" in sht2["J1"].value: #（所長的情况）
                        if sht["J" + str(col.row)].value is not None and "所長" in sht["J" + str(col.row)].value: #（如果这边也有所長）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "所長" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value: #（如果这边有职位 但不是所長）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    elif "センター長" in sht2["J1"].value: #（センター長的情况）
                        if sht["J" + str(col.row)].value is not None and "センター長" in sht["J" + str(col.row)].value: #（如果这边也有センター長）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "センター長" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value: #（如果这边有职位 但没有センター長和支店長）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    elif "チーフ" in sht2["J1"].value: #（チーフ的情况）
                        if sht["J" + str(col.row)].value is not None and "チーフ" in sht["J" + str(col.row)].value: #（如果这边也有チーフ）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "チーフ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value: #（如果这边有职位 也没有チーフ和センター長和支店長）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監察役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    elif "主任" in sht2["J1"].value: #（主任的情况）
                        if sht["J" + str(col.row)].value is not None and "主任" in sht["J" + str(col.row)].value: #（如果这边也有主任）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "主任" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value: #（如果这边有职位 也没有主任和チーフ和センター長和支店長）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    # 本部的四个职位的判断
                    elif "本部長" in sht2["J1"].value: #（本部长的情况）
                        if sht["J" + str(col.row)].value is not None and "本部長" in sht["J" + str(col.row)].value: #（如果这边也有本部长）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "本部長" not in sht["J" + str(col.row)].value: #（如果有职位 但不是本部长）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break


                    elif "部長" in sht2["J1"].value and "本部長" not in sht2["J1"].value: #（部长的情况）
                        if (sht["J" + str(col.row)].value is not None) and ("部長" in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value): #（如果这边也有部长）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "部長" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value: #（如果有职位 但不是（本）部长）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    elif "マネージャ" in sht2["J1"].value: #（マネージャ的情况）
                        if (sht["J" + str(col.row)].value is not None) and "マネージャ" in sht["J" + str(col.row)].value: #（如果这边也有マネージャ）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value: #（如果有职位 但不是（本）部长）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    elif "リーダ" in sht2["J1"].value: #（リーダ的情况）
                        if (sht["J" + str(col.row)].value is not None) and "リーダ" in sht["J" + str(col.row)].value: #（如果这边也有リーダ）
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                        elif sht["J" + str(col.row)].value is not None and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value: #（如果有职位 但不是（本）部长）
                            break

                        #（如果这边没有任何职位）
                        elif (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            break

                    #（不是空值 没有职位但有驻在）
                    if "監察役" not in sht2["J1"].value and "取締役" not in sht2["J1"].value and "支社長" not in sht2["J1"].value and "本部長" not in sht2["J1"].value and "部長" not in sht2["J1"].value and "マネージャ" not in sht2["J1"].value and "センター長" not in sht2["J1"].value and "チーフ" not in sht2["J1"].value and "主任" not in sht2["J1"].value and "リーダ" not in sht2["J1"].value and "支店長" not in sht2["J1"].value:
                        if (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                elif sht2["J1"].value is None: #（无职位 空值的情况）
                    print("进入1-2分支")

                    #（这边也是无职位的情况）
                    if (sht["J" + str(col.row)].value is None) or ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                        if sht["E" + str(col.row)].value is not None:
                            print("进入1-2-1分支")
                            compare_list.append(sht["E" + str(col.row)].value)
                            old_list.append(sht["E" + str(col.row)].value)

                    # （这边有驻在情况）
                    elif (sht["J" + str(col.row)].value is not None) and ("監査役" not in sht["J" + str(col.row)].value and "取締役" not in sht["J" + str(col.row)].value and "支社長" not in sht["J" + str(col.row)].value and "本部長" not in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row)].value and "マネージャ" not in sht["J" + str(col.row)].value and "センター長" not in sht["J" + str(col.row)].value and "チーフ" not in sht["J" + str(col.row)].value and "主任" not in sht["J" + str(col.row)].value and "リーダ" not in sht["J" + str(col.row)].value and "支店長" not in sht["J" + str(col.row)].value):
                        print("进入1-2-2分支")
                        compare_list.append(sht["E" + str(col.row)].value)
                        old_list.append(sht["E" + str(col.row)].value)

                    #（这边有职位的情况）
                    if (sht["J" + str(col.row)].value is not None) and ("監察役" in sht["J" + str(col.row)].value or "取締役" in sht["J" + str(col.row)].value or "支社長" in sht["J" + str(col.row)].value or "本部長" in sht["J" + str(col.row)].value or "部長" in sht["J" + str(col.row)].value or "マネージャ" in sht["J" + str(col.row)].value or "センター長" in sht["J" + str(col.row)].value or "チーフ" in sht["J" + str(col.row)].value or "主任" in sht["J" + str(col.row)].value or "リーダ" in sht["J" + str(col.row)].value or "支店長" in sht["J" + str(col.row)].value or "所長" in sht["J" + str(col.row)].value):
                        print("进入1-2-3分支")
                        k += 1
                        if k == same_busho_kojin:
                            lastone += 1
                            break

    #没有同部署的情况下
    elif same_busho == 0:
        print("进入2号分支")

        lastone = sht.max_row + 1

        for each_cell in sht["H"]:
            if each_cell.value.startswith("ad") and not each_cell.value.startswith("ads") and not each_cell.value.startswith("ada"):
                busho_code = int(sht2["H1"].value.split("@")[0][2:])
                each_busho = (int(each_cell.value.split("@")[0][2:]))

                if each_busho > busho_code:
                    lastone = each_cell.row
                    break

            elif not each_cell.value.startswith("ad") and not each_cell.value.startswith("ads") and not each_cell.value.startswith("ada"):
                lastone = sht.max_row + 1



    print(f"插入的行号为:{lastone}")

    compare_list.append(sht2["E1"].value)
    old_list.append(sht2["E1"].value)
    
    #到这里为止 这两个列表应该是相同的

    compare_list.sort()

    print(f"排序前：{old_list}")
    print(f"排序后：{compare_list}")


    def insert_to_excel(row):

        sht.insert_rows(row)

        i = 0

        for col in sht.iter_cols(min_row=row, max_col=22, max_row=row):
             for cell in col:
                 cell.value = temps[i]
                 i += 1

        sht2.delete_rows(1)

        return

    if len(compare_list) == 1:

        temps = []

        for cell in sht2[1]:
            temps.append(cell.value)

        print(temps)

        insert_to_excel(lastone)

    elif len(compare_list) != 1:

        if old_list != compare_list:
            print(compare_list.index(sht2["E1"].value) + 1) #取得排序后列表中 排序目标的下一个人的index
            print(compare_list[compare_list.index(sht2["E1"].value) + 1]) #通过index取得下一个人的人名

            for t in range(lastone,2,-1):
                if sht["E" + str(t)].value == compare_list[compare_list.index(sht2["E1"].value) + 1]:
                    print(t) #应该插入的位置 行数
                    break

            temps = []

            for cell in sht2[1]:
                temps.append(cell.value)

            print(temps)

            insert_to_excel(t)

        elif old_list == compare_list:
            print(compare_list.index(sht2["E1"].value) - 1) #取得排序后列表中 排序目标的上一个人的index
            print(compare_list[compare_list.index(sht2["E1"].value) - 1]) #通过index取得上一个人的人名

            for t in range(lastone,2,-1):
                if sht["E" + str(t)].value == compare_list[compare_list.index(sht2["E1"].value) - 1]:
                    print(t + 1) #应该插入的位置 行数
                    break

            temps = []

            for cell in sht2[1]:
                temps.append(cell.value)

            print(temps)

            insert_to_excel(t + 1)


book.save(f"{filename}_changed.xlsx")

msgbox("完成")