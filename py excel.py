import openpyxl

book = openpyxl.load_workbook("20210309_Rakumo更新用ファイル.xlsx")

sht = book.get_sheet_by_name("Sheet1")
sht2 = book.get_sheet_by_name("Sheet2")

colV = sht["V"]

targets = []

for col in colV:
    if col.value == "変更":
        RowNumber = col.row
        BuSho = sht["H" + str(RowNumber)].value

for row in sht.iter_rows(min_row=2):
    if row[21].value is None:
        continue
    values = []
    for col in row:
        values.append(col.value)

    print(values)
    targets.append(values)

print(f"共有{len(targets)}条数据需要排列")

for data in targets:
    sht2.append(data)

for col in colV:
    if col.value == "変更":
        RowNumber = col.row
        sht.delete_rows(RowNumber)
# 以上 把第一页中需要排列的数据剪切到第二页

print(sht2.max_row)



print(sht2["H1"].value)

colH = sht["H"]
same_busho = 0

compare_list = []
old_list = []

for col in colH:
    if col.value == sht2["H1"].value:
        print(col.row)
        same_busho += 1
        lastone = col.row #同部署的最后一条的行数

print(f"在Sheet1中,共有{same_busho}个相同部署的,最后一个的行号是{lastone}")

for col in colH:
    if col.value == sht2["H1"].value and "mbox" not in sht["A" + str(col.row)].value and "atenda" not in sht["A" + str(col.row)].value: #个人邮箱的判断
        lastone = col.row
        if sht["J" + str(col.row)].value == sht2["J1"].value: #职位的判断（无职位的情况）
            #lastone = col.row
            compare_list.append(sht["E" + str(col.row)].value)
            old_list.append(sht["E" + str(col.row)].value)
        elif "部長" in sht2["J1"]:
            if "支社長" in sht["J" + str(col.row)].value and "部長" not in sht["J" + str(col.row + 1)].value:
                # compare_list.append(sht["E" + str(col.row)].value)
                # old_list.append(sht["E" + str(col.row)].value)

compare_list.append(sht2["E1"].value)
old_list.append(sht2["E1"].value)

print(f"同部署个人邮箱最后一行的行号是{lastone}")

compare_list.sort()

print(f"排序前：{old_list}")
print(f"排序后：{compare_list}")

print(compare_list.index(sht2["E1"].value) + 1)

print(compare_list[compare_list.index(sht2["E1"].value) + 1])

for t in range(col.row,2,-1):
    if sht["E" + str(t)].value == compare_list[compare_list.index(sht2["E1"].value) + 1]:
        print(t) #应该插入的位置 行数
        break

temps = []

for cell in sht2[1]:
    temps.append(cell.value)

print(temps)

sht.insert_rows(t)

i = 0

for col in sht.iter_cols(min_row=t, max_col=22, max_row=t):
     for cell in col:
         cell.value = temps[i]
         i += 1

sht2.delete_rows(1)

book.save("changed.xlsx")
